"""utils functions"""
import pathlib
import pandas as pd
import pyexcel as p
import win32com.client as win32
import xlrd
from pathlib import Path
from openpyxl.workbook import Workbook

COLUMNS = ("DATA", "TIPO", "DESCRIZIONE", "IN ENTRATA/IN USCITA", "SALDO")
VALUTE_MAPPING = {
    "USD":"$",
    "EUR":"€"
}

def get_valuta_symbol(acronym:str) -> str:
    return VALUTE_MAPPING.get(acronym, None)

def convert_to_csv_text(vals:list[list[str]]) -> str:
    text_vals = "\n".join([";".join(v) for v in vals])
    return text_vals
    
def convert_xls_to_xlsx(path: Path) -> Path:
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(path.absolute())
    # FileFormat=51 is for .xlsx extension
    new_path_file = str(path.absolute().with_suffix(".xlsx")) 
    wb.SaveAs(new_path_file, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    new_path = pathlib.Path(new_path_file)
    path.unlink()
    return new_path

def convert_xls_to_xlsx2(path:Path) ->Path:
    path_converted = path.parent / (path.name + "x") 
    df = pd.read_excel(path.absolute())
    # Save as .xlsx
    df.to_excel(path_converted, index=False)
    path.unlink()
    return path_converted

def convert_xls_to_xlsx3(path:Path) ->Path:
    path_converted = path.parent / (path.name + "x") 
    p.save_book_as(file_name=path.name,
               dest_file_name=path_converted.name)
    path.unlink()
    return path_converted

def convert_xls_to_xlsx4(path:Path) -> Path:
    path_converted = path.parent / (path.name + "x")     
    book_xls = xlrd.open_workbook(path.absolute())
    book_xlsx = Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    
    book_xlsx.save(path_converted.absolute())
    path.unlink()
    return path_converted

def convert_b_format_month(data:str) -> str:
    """converts the string data like '01 ago 2024' to '01/08/2024'"""
    ita_b = ("gen", "feb", "mar", "apr", "mag", "giu", "lug", "ago", "set", "ott", "nov", "dic")
    if len(data) != 11: raise Exception(f"Can't process provided string {data}")
    month = data[3:6]
    n_month = str(ita_b.index(month) + 1)
    data = f"{data[0:2]}/{ '0' + n_month if len(n_month) == 1 else n_month}/{data[7:]}"
    return data