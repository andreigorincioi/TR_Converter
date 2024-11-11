"""This library is used to convert data from xls file to csv"""
import datetime
from configparser import ConfigParser
import pathlib
import openpyxl as xl
from logger import LoggerModule
from utils import convert_to_csv_text, convert_xls_to_xlsx2, COLUMNS, convert_xls_to_xlsx3, convert_xls_to_xlsx4

MAPPING_XLS = {
    "DATA":"A",
    "Data Valuta":"B",
    "IN ENTRATA/IN USCITA":"C",
    "VALUTA":"D",
    "DESCRIZIONE":"E",
    "TIPO":"F"}
cwd = pathlib.Path.cwd()
FROM_FOLDER = cwd / 'ToConvert'  
TO_FOLDER = cwd / 'Converted'
logger = LoggerModule('ExtractFromTRRPdf') 

try:
    cnf = ConfigParser()
    cnf.read("config.ini")
except Exception as ex:
    logger.error('config.ini not found', ex)        

try:
    DECIMAL_TYPE = int(cnf['general']['decimal_type'])
except Exception as ex:
    logger.error('config-log.ini')

def main():
    for path in FROM_FOLDER.iterdir():
        if path.name.lower().find("webank") == -1: continue
        vals = process_file(path)
        save_to_csv(path, vals)

def process_file(path_file:pathlib.Path):
    file_name = path_file.name
    file_type = file_name[file_name.find("."):]
    if file_type == ".xls": path_file = convert_xls_to_xlsx4(path_file)
    wb = xl.open(path_file.absolute(), read_only=True, data_only=True)
    ws = wb.active
    vals = [list(COLUMNS)]
    for row_number in range(2, ws.max_row):
        line_vals = []
        for col in COLUMNS:
            match col:
                case "DATA"|"DESCRIZIONE"|"TIPO":
                    cell_pos = f"{MAPPING_XLS.get(col)}{row_number}"
                    value = ws[cell_pos].value
                    if type(value) == datetime.datetime: value = value.strftime("%d/%m/%Y")
                    if value is None: value = ''
                    if type(value) != str: raise Exception("Value is not string") 
                    line_vals.append(value)
                    # TODO add decimal point handling based on confi ini variable
                case "IN ENTRATA/IN USCITA":
                    cell_pos = f"{MAPPING_XLS.get(col)}{row_number}"
                    value = ws[cell_pos].value
                    line_vals.append(value)
        line_vals.append("-")
        vals.append(line_vals)
    return vals

def save_to_csv(path:pathlib.Path, values:list[list[str]]):
    text_vals = convert_to_csv_text(values)
    file_name = path.name.removesuffix(".xls") + ".csv"
    file_path = TO_FOLDER / file_name
    with file_path.open("wt", encoding="utf-8") as fw:
        fw.write(text_vals)

if __name__ == "__main__":
    main()
