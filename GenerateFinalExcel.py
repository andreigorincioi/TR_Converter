"""Genearates the final excel file \
    from the csv files of the Converted folder"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CSV_COLUMNS = ("DATA", "TIPO", "DESCRIZIONE", "IN ENTRATA/IN USCITA", "SALDO")

cwd = Path().cwd()
FOLDER_TO = cwd / "FinalExcel"
FOLDER_FROM_CSV = cwd / "Converted"

def convert_csv_to_excel_ws(ws: Worksheet, csv_reader) -> int:
    """imports csv data in the excel page"""
    for line in csv_reader:
        ws.append(line)
    return csv_reader.line_num

def create_tot_ws(ws_tot:Worksheet ,ws_names: list[str], ws_len:list[str]):
    n_pages = len(ws_names)
    for i in range(n_pages):
        ws_tot.cell(1 + i, 1).value = f"Total {ws_names[i]}"
        ws_tot.cell(1 + i, 2).value = f"={ws_names[i]}!{get_column_letter(i+1)}"
    ws_tot.cell(1, n_pages + 1).value = "Total of totals"
    ws_tot.cell(2, n_pages + 1).value = f"={''.join(ws_names)}"

def main(date: str):
    wb = Workbook()
    ws_len: list[int] = []
    ws_names: list[str] = []
    csv_paths = [path for path in FOLDER_FROM_CSV.iterdir() 
                    if path.name.endswith(".csv") and path.name.rfind(date) != -1]
    for csv_path in csv_paths:
        with csv_path.open("r") as fr: 
            csv_reader = csv.reader(fr.read(), delimiter=";")
            name_page = csv_path.name.removesuffix(".csv")
            ws_names.append(name_page)
            ws = wb.create_sheet(name_page)
            len_page = convert_csv_to_excel_ws(ws,csv_reader)
            ws_len.append(len_page)
    ws_tot: Worksheet = wb.create_sheet("Totali")
    create_tot_ws(ws_tot, ws_names, ws_len)
    path_excel = (cwd / f"FinalExcel\\{date}_Totals.xlsx")
    wb.save(path_excel)

if __name__ == '__main__':
    main("09_2024")
