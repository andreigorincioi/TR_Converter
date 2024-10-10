import enum
import openpyxl


def create_workbook_from_tr_converter_data(values:list[list[str]], date:str):
    w = openpyxl.Workbook()
    w.create_sheet(f"EstrattoConto_{date}", 0)
    sheet = w.get_sheet_by_name(f"EstrattoConto_{date}")
    # write data
    for row, l_val in enumerate(values):
        for col, val in enumerate(l_val):
            sheet.cell(row + 1, col + 1).value = val
    
    w.save(f"Converted\\TRR_{date}.xlsx")